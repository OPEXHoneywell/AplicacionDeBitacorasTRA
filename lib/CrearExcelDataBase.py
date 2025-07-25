import customtkinter
from tkcalendar import Calendar
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pyodbc

# Rutas de imágenes (ajusta si es necesario)
image_paths = [
    "PruebasDeFramesIMG/imgTrabajoCaliente.png", 
    "PruebasDeFramesIMG/imgConfinados.png",
    "PruebasDeFramesIMG/imgAlturas.png",
    "PruebasDeFramesIMG/imgBloqueo.png",
    "PruebasDeFramesIMG/imgEnergia.png",
    "PruebasDeFramesIMG/imgTuberias.png",
    "PruebasDeFramesIMG/imgAdmCambios.png",
    "PruebasDeFramesIMG/imgExcavaciones.png",
    "PruebasDeFramesIMG/imgLevantamiento.png",
    "PruebasDeFramesIMG/imgRepAb.png"  # Para TRA Activo
]

def generar_excel_desde_db(fecha_inicio, fecha_fin, ruta_guardado=None):
    try:
        # Ahora se espera que fecha_inicio y fecha_fin estén en formato YYYY-MM-DD (DATE)
        # Si vienen en otro formato, intentar convertir
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
        except Exception:
            # Intentar con otros formatos comunes
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, "%d/%m/%Y")
                fecha_fin_dt = datetime.strptime(fecha_fin, "%d/%m/%Y")
            except Exception:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, "%m_%d_%Y")
                fecha_fin_dt = datetime.strptime(fecha_fin, "%m_%d_%Y")
        fecha_inicio_sql = fecha_inicio_dt.strftime("%Y-%m-%d")
        fecha_fin_sql = fecha_fin_dt.strftime("%Y-%m-%d")
        # Conexión a SQL Server
        conexion = pyodbc.connect(
            'DRIVER={ODBC Driver 17 for SQL Server};'
            'SERVER=MX29W1009;'
            'DATABASE=DB_BitacoraTRA;'
            'UID=PracticanteHSE;'
            'PWD=PracticanteHSE2025;'
        )
        cursor = conexion.cursor()
        # Consulta todos los registros en el rango de fechas usando tipo DATE
        cursor.execute('''
            SELECT tra_id, fecha, edificio, area, lugar, hora, descripcion, compania, responsable,
                   trabajo_caliente, espacios_confinados, trabajo_en_alturas, bloqueo_etiquetado_prueba,
                   energia_en_vivo, interrupcion_de_tuberia, administracion_del_cambio, excavaciones,
                   equipo_de_levantamiento, tra_activo, seleccion_boton_matriz, firmado_por, estado, inspector, razon
            FROM registros_tra
            WHERE fecha >= ? AND fecha <= ?
            ORDER BY tra_id
        ''', (fecha_inicio_sql, fecha_fin_sql))
        filas = []
        filas_auditados = []
        filas_cancelados = []
        for row in cursor.fetchall():
            filas.append(row)
            if str(row[21]).strip().upper() == "AUDITADO":
                filas_auditados.append(row)
            if str(row[21]).strip().upper() == "CANCELADO":
                filas_cancelados.append(row)
        if not filas:
            cursor.close()
            conexion.close()
            messagebox.showinfo("Sin datos", f"No hay registros entre {fecha_inicio} y {fecha_fin}")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial TRA"
        headers = [
            "ID", "Fecha", "Edificio", "Area", "Lugar", "Hora", "Descripcion", "Compania", "Responsable",
            "Trabajo Caliente", "Espacios Confinados", "Trabajo en alturas", "Bloqueo/Etiquetado/Prueba",
            "Energia en vivo", "Interrupcion de tuberia", "Administracion del cambio", "Excavaciones",
            "Equipo de Levantamiento", "TRA Activo", "Grado de riesgo", "Firmado por", "Estado", "Inspector", "Razón cancelación"
        ]
        ws.append(headers)
        for col_num, img_path in enumerate(image_paths, start=10):
            if os.path.exists(img_path):
                img = ExcelImage(img_path)
                img.width = 30
                img.height = 30
                col_letter = get_column_letter(col_num)
                cell_coord = f"{col_letter}1"
                ws.add_image(img, cell_coord)
                ws.column_dimensions[col_letter].width = 12
        ws.row_dimensions[1].height = 35
        for row_idx, row in enumerate(filas, start=2):
            ws.append(list(row))
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(max_length + 2, 15)  # mínimo 15 para buena visualización
        # Crear tabla para habilitar filtros y ordenamiento en Excel
        from openpyxl.worksheet.table import Table, TableStyleInfo
        table = Table(displayName="HistorialTRA", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)
        # Colorea checks en verde y limpia valores 1/0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=10, max_col=20):
            for col_idx, cell in enumerate(row, start=10):
                # Checkboxes (columnas 10-18)
                if 10 <= col_idx <= 18:
                    if cell.value == 1:
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                        cell.value = ""
                    elif cell.value == 0:
                        cell.value = ""
                # Celda de color de la matriz (columna 20)
                if col_idx == 20:
                    color_matriz = cell.value
                    if color_matriz and isinstance(color_matriz, str) and color_matriz.strip():
                        color_hex = color_matriz.strip().replace('#', '')
                        if len(color_hex) == 6:
                            color_hex = 'FF' + color_hex  # openpyxl espera ARGB, FF=opaco
                        elif len(color_hex) == 8:
                            pass  # asume ARGB
                        else:
                            color_hex = 'FF00FF00'  # fallback verde opaco
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                        cell.value = ""  # Limpiar el valor de la celda para que no se muestre el hexcode
                # Celda de selección de matriz (columna 19)
                if col_idx == 19:
                    # Si el valor es 1, pintamos de verde y limpiamos el valor
                    if cell.value == 1:
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                        cell.value = ""
                    elif cell.value == 0:
                        cell.value = ""
        # Segunda hoja: solo TRA auditados
        headers_auditados = [
            "TRA", "Fecha", "Edificio", "Area", "Lugar", "Hora", "Inspector", "Grado de riesgo",
            "Trabajo Caliente", "Espacios Confinados", "Trabajo en alturas", "Bloqueo/Etiquetado/Prueba",
            "Energia en vivo", "Interrupcion de tuberia", "Administracion del cambio", "Excavaciones",
            "Equipo de Levantamiento", "TRA Activo", "Firmado por", "Estado", "Razón cancelación"
        ]
        ws_auditados = wb.create_sheet(title="TRA Auditados")
        ws_auditados.append(headers_auditados)
        for col_num, img_path in enumerate(image_paths, start=9):
            if os.path.exists(img_path):
                img = ExcelImage(img_path)
                img.width = 30
                img.height = 30
                col_letter = get_column_letter(col_num)
                cell_coord = f"{col_letter}1"
                ws_auditados.add_image(img, cell_coord)
                ws_auditados.column_dimensions[col_letter].width = 12
        ws_auditados.row_dimensions[1].height = 35
        for row in filas_auditados:
            nueva_row = [
                row[0],  # TRA (id)
                row[1],  # Fecha
                row[2],  # Edificio
                row[3],  # Area
                row[4],  # Lugar
                row[5],  # Hora
                row[22], # Inspector
                row[19], # Seleccion Matriz (Grado de riesgo)
                row[9],  # Trabajo Caliente
                row[10], # Espacios Confinados
                row[11], # Trabajo en alturas
                row[12], # Bloqueo/Etiquetado/Prueba
                row[13], # Energia en vivo
                row[14], # Interrupcion de tuberia
                row[15], # Administracion del cambio
                row[16], # Excavaciones
                row[17], # Equipo de Levantamiento
                row[18], # TRA Activo
                row[20], # Firmado por
                row[21], # Estado
                row[23] if len(row) > 23 else ""  # Razón cancelación
            ]
            ws_auditados.append(nueva_row)
        # Formato de checks y colores
        for row in ws_auditados.iter_rows(min_row=2, max_row=ws_auditados.max_row, min_col=9, max_col=18):
            for col_idx, cell in enumerate(row, start=9):
                if 9 <= col_idx <= 18:  # Incluye TRA Activo
                    if cell.value == 1:
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                        cell.value = ""
                    elif cell.value == 0:
                        cell.value = ""
        # Grado de riesgo (col 8)
        for row in ws_auditados.iter_rows(min_row=2, max_row=ws_auditados.max_row, min_col=8, max_col=8):
            for cell in row:
                color_matriz = cell.value
                if color_matriz and isinstance(color_matriz, str) and color_matriz.strip():
                    color_hex = color_matriz.strip().replace('#', '')
                    if len(color_hex) == 6:
                        color_hex = 'FF' + color_hex
                    elif len(color_hex) == 8:
                        pass
                    else:
                        color_hex = 'FF00FF00'
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    cell.value = ""  # Limpiar el valor de la celda para que no se muestre el hexcode
        # Ajustar alineación de los encabezados al centro
        from openpyxl.styles import Alignment
        for cell in ws_auditados[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, ws_auditados.max_column + 1):
            col_letter = get_column_letter(col)
            ws_auditados.column_dimensions[col_letter].width = 20
        
         # Crear tabla para habilitar filtros y ordenamiento en Excel
        from openpyxl.worksheet.table import Table, TableStyleInfo
        if ws_auditados.max_row > 1:  # Solo crear tabla si hay datos
            # Crear tabla para TRA Auditados
            table_auditados = Table(displayName="TRA_Auditados", ref=f"A1:{get_column_letter(ws_auditados.max_column)}{ws_auditados.max_row}")
            style_auditados = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table_auditados.tableStyleInfo = style_auditados
            ws_auditados.add_table(table_auditados)
        
        
        # --- Hoja 3: TRA Cancelados ---
        ws_cancelados = wb.create_sheet(title="TRA Cancelados")
        headers_cancelados = [
            "Fecha", "TRA ID", "Edificio", "Area", "Lugar", "Hora", "Cancelado por", "Razón de cancelación"
        ]
        ws_cancelados.append(headers_cancelados)
        for row in filas_cancelados:
            nueva_fila = [
                row[1],  # Fecha
                row[0],  # TRA ID
                row[2],  # Edificio
                row[3],  # Area
                row[4],  # Lugar
                row[5],  # Hora
                row[22], # Cancelado por (inspector)
                row[23] if len(row) > 23 else ""  # Razón de cancelación
            ]
            ws_cancelados.append(nueva_fila)
        if ws_cancelados.max_row > 1:  # Solo crear tabla si hay datos
            # Crear tabla para habilitar filtros y ordenamiento en Excel
            table_cancelados = Table(
            displayName="TRA_Cancelados",
            ref=f"A1:{get_column_letter(ws_cancelados.max_column)}{ws_cancelados.max_row}"
            )
            style_cancelados = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            table_cancelados.tableStyleInfo = style_cancelados
            ws_cancelados.add_table(table_cancelados)
        cursor.close()
        conexion.close()
        if ruta_guardado:
            excel_filename = ruta_guardado
        else:
            excel_filename = f"datos_TRA_{fecha_inicio}_a_{fecha_fin}.xlsx"
        wb.save(excel_filename)
        messagebox.showinfo("Éxito", f"Archivo Excel generado exitosamente como '{excel_filename}'")
        return excel_filename
    except Exception as e:
        messagebox.showerror("Error", f"Error al generar Excel: {str(e)}")
        return None

def app(master=None):
    if master is None:
        root = customtkinter.CTk()
        root.title("Generar Excel desde Base de Datos por Rango de Fechas")
        container = root
    else:
        container = master
        for widget in container.winfo_children():
            widget.destroy()
    frame = customtkinter.CTkFrame(container)
    frame.pack(padx=20, pady=20, fill="both", expand=True)

    fecha_inicio_var = tk.StringVar()
    fecha_fin_var = tk.StringVar()

    def seleccionar_fecha(var, label):
        popup = tk.Toplevel(container)
        popup.title("Seleccionar Fecha")
        cal = Calendar(popup, selectmode='day')
        cal.pack(pady=20)
        def confirmar_fecha():
            fecha_obj = cal.selection_get()
            try:
                fecha_str = fecha_obj.strftime("%#m_%#d_%Y")
            except:
                fecha_str = fecha_obj.strftime("%m_%d_%Y").lstrip("0").replace("_0", "_")
            var.set(fecha_str)
            label.configure(text=f"{label.cget('text').split(':')[0]}: {fecha_str}")
            popup.destroy()
        boton = customtkinter.CTkButton(popup, text="Seleccionar", command=confirmar_fecha)
        boton.pack(pady=10)

    label_inicio = customtkinter.CTkLabel(frame, text="Fecha inicio: No seleccionada")
    label_inicio.pack(pady=5)
    boton_inicio = customtkinter.CTkButton(frame, text="Seleccionar Fecha Inicio", command=lambda: seleccionar_fecha(fecha_inicio_var, label_inicio))
    boton_inicio.pack(pady=5)

    label_fin = customtkinter.CTkLabel(frame, text="Fecha fin: No seleccionada")
    label_fin.pack(pady=5)
    boton_fin = customtkinter.CTkButton(frame, text="Seleccionar Fecha Fin", command=lambda: seleccionar_fecha(fecha_fin_var, label_fin))
    boton_fin.pack(pady=5)

    def generar_y_mostrar():
        fecha_inicio = fecha_inicio_var.get()
        fecha_fin = fecha_fin_var.get()
        if not fecha_inicio or not fecha_fin:
            messagebox.showerror("Error", "Debes seleccionar ambas fechas.")
            return
        # Preguntar al usuario dónde guardar el archivo
        default_filename = f"datos_TRA_{fecha_inicio}_a_{fecha_fin}.xlsx"
        ruta_guardado = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            initialfile=default_filename,
            title="Guardar Excel como..."
        )
        if not ruta_guardado:
            return  # Usuario canceló
        excel_filename = generar_excel_desde_db(fecha_inicio, fecha_fin, ruta_guardado)
        if excel_filename:
            mostrar_boton_abrir_excel(excel_filename)

    boton_generar = customtkinter.CTkButton(
        frame,
        text="Generar Excel del Rango",
        command=generar_y_mostrar,
        fg_color="green",
        hover_color="darkgreen"
    )
    boton_generar.pack(pady=20)

    def mostrar_boton_abrir_excel(excel_filename):
        # Elimina cualquier botón anterior de abrir excel
        for widget in frame.winfo_children():
            if getattr(widget, 'is_abrir_excel', False):
                widget.destroy()
        def abrir_excel():
            abs_path = os.path.abspath(excel_filename)
            if not os.path.exists(abs_path):
                messagebox.showerror("Error", f"No se encontró el archivo: {abs_path}")
                return
            try:
                os.startfile(abs_path)
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el archivo: {str(e)}")
        # Cargar imagen de Excel si existe
        try:
            from PIL import Image
            img = Image.open("PruebasDeFramesIMG/excel_icon.png")
            img = img.convert("RGBA")  # Asegura canal alfa
            excel_img = customtkinter.CTkImage(light_image=img, size=(32,32))
        except Exception as e:
            excel_img = None
        boton_abrir = customtkinter.CTkButton(
            frame,
            text="Abrir en Excel",
            image=excel_img if excel_img else None,
            compound="left" if excel_img else None,
            command=abrir_excel,
            fg_color="#217346",
            hover_color="#185c37"
        )
        boton_abrir.is_abrir_excel = True
        boton_abrir.pack(pady=10)

    if master is None:
        container.mainloop()

if __name__ == "__main__":
    
    app()
